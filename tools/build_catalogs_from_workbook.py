from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / 'data'
SEED_XLSX = DATA_DIR / 'pv_inverter_catalog_seed_2026-04-09.xlsx'
EXPANDED_XLSX = DATA_DIR / 'pv_inverter_catalog_expanded_models_2026-04-09.xlsx'
OUT_PV = DATA_DIR / 'pv-modules.json'
OUT_GRID = DATA_DIR / 'grid-inverters.json'
OUT_HYBRID = DATA_DIR / 'hybrid-inverters.json'


def clean_text(v: Any, max_len: int = 240) -> str:
    s = '' if v is None else str(v)
    s = re.sub(r'[\x00-\x1F\x7F]+', '', s).strip()
    return s[:max_len]


def slug_id(*parts: str) -> str:
    s = '_'.join(clean_text(p, 80).lower() for p in parts if clean_text(p, 80))
    s = re.sub(r'[^a-z0-9_]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s[:96]


def as_float(v: Any) -> Optional[float]:
    try:
        n = float(v)
    except Exception:
        return None
    if math.isfinite(n):
        return n
    return None


def read_sheet_rows(path: Path, name: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[name]
    headers = [clean_text(c.value, 120) for c in ws[1]]
    out: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(x is not None and str(x).strip() for x in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        out.append(item)
    return out


def round_n(v: float, n: int) -> float:
    return float(f"{v:.{n}f}")


def interp(min_v: float, max_v: float, pmin: float, pmax: float, p: float) -> float:
    if abs(pmax - pmin) < 1e-9:
        return float(min_v)
    t = (p - pmin) / (pmax - pmin)
    return min_v + t * (max_v - min_v)


def parse_module_watt(model: str) -> Optional[float]:
    s = clean_text(model, 180).upper()
    m = re.search(r'-(\d{3})(?:[A-Z/]|$)', s)
    if m:
        return float(m.group(1))
    m = re.search(r'(\d{3})(?:W|M|N)\b', s)
    if m:
        return float(m.group(1))
    return None


def parse_inverter_kw(model: str, manufacturer: str) -> Optional[float]:
    s = clean_text(model, 220)
    su = s.upper().replace(' ', '')

    patterns = [
        r'(\d+(?:\.\d+)?)K(?:TL|TLM|TLX|SG|G|$)',
        r'(\d+(?:\.\d+)?)K\b',
        r'-(\d+(?:\.\d+)?)-TL',
        r'PVS-(\d+(?:\.\d+)?)',
        r'TRIO-(\d+(?:\.\d+)?)',
        r'BLUEPLANET(\d+(?:\.\d+)?)',
        r'(?<!\d)(\d{3,4})(?=TL)',
    ]
    for pat in patterns:
        m = re.search(pat, su)
        if not m:
            continue
        val = float(m.group(1))
        if val >= 1000:
            val = val / 1000.0
        if 0.2 <= val <= 500:
            return val

    if clean_text(manufacturer).lower().startswith('fronius'):
        m = re.search(r'\b(\d+(?:\.\d+)?)-\d', s)
        if m:
            val = float(m.group(1))
            if 0.2 <= val <= 500:
                return val

    nums = [float(x) for x in re.findall(r'\d+(?:\.\d+)?', su)]
    candidates: List[float] = []
    for n in nums:
        x = n / 1000.0 if n >= 1000 else n
        if 0.2 <= x <= 500:
            candidates.append(x)
    if candidates:
        return max(candidates)

    return None


def coeff_defaults(manufacturer: str) -> Tuple[float, float, float]:
    m = clean_text(manufacturer).lower()
    if 'ja solar' in m:
        return -0.0028, 0.00045, -0.0035
    if 'risen' in m:
        return -0.0026, 0.00048, -0.0034
    if 'longi' in m:
        return -0.0026, 0.00046, -0.0030
    if 'jinko' in m:
        return -0.0026, 0.00048, -0.0030
    return -0.0026, 0.00048, -0.0032


def cells_guess(model: str) -> int:
    u = clean_text(model).upper()
    m = re.search(r'RSM(\d{3})', u)
    if m:
        return int(m.group(1))
    if 'JAM66' in u:
        return 66
    if 'JAM72' in u or 'JKM' in u or 'LR5-72' in u:
        return 72
    for n in (54, 60, 66, 72, 78, 108, 110, 120, 132, 144, 150, 156):
        if str(n) in u:
            return n
    return 72


def base_seed_family_specs(seed_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for r in seed_rows:
        pmin = as_float(r.get('pmax_w_min'))
        pmax = as_float(r.get('pmax_w_max'))
        voc_min = as_float(r.get('voc_v_min'))
        voc_max = as_float(r.get('voc_v_max'))
        vmp_min = as_float(r.get('vmp_v_min'))
        vmp_max = as_float(r.get('vmp_v_max'))
        isc_min = as_float(r.get('isc_a_min'))
        isc_max = as_float(r.get('isc_a_max'))
        imp_min = as_float(r.get('imp_a_min'))
        imp_max = as_float(r.get('imp_a_max'))
        if None in (pmin, pmax, voc_min, voc_max, vmp_min, vmp_max, isc_min, isc_max, imp_min, imp_max):
            continue
        key = f"{clean_text(r.get('manufacturer'))}|{clean_text(r.get('model_family'))}"
        out[key] = {
            'kind': 'datasheet-range',
            'pmin': pmin,
            'pmax': pmax,
            'voc_min': voc_min,
            'voc_max': voc_max,
            'vmp_min': vmp_min,
            'vmp_max': vmp_max,
            'isc_min': isc_min,
            'isc_max': isc_max,
            'imp_min': imp_min,
            'imp_max': imp_max,
            'maxSystemV': int(as_float(r.get('max_system_vdc')) or 1500),
            'source_url': clean_text(r.get('source_url'), 300),
            'source_note': clean_text(r.get('source_note'), 300),
        }

    # Extend Jinko family to full 550-590 range using slope from official 550-570 data
    jinko_key = 'Jinko Solar|JKM550-570M-72HL4-(V)'
    if jinko_key in out:
        js = dict(out[jinko_key])
        pmin = js['pmin']
        pmax = js['pmax']
        span = pmax - pmin if pmax > pmin else 20.0
        ext_to = 590.0
        if ext_to > pmax:
            f = (ext_to - pmin) / span
            js['pmax'] = ext_to
            js['voc_max'] = js['voc_min'] + (js['voc_max'] - js['voc_min']) * f
            js['vmp_max'] = js['vmp_min'] + (js['vmp_max'] - js['vmp_min']) * f
            js['isc_max'] = js['isc_min'] + (js['isc_max'] - js['isc_min']) * f
            js['imp_max'] = js['imp_min'] + (js['imp_max'] - js['imp_min']) * f
            js['source_note'] = clean_text((js.get('source_note') or '') + ' Extended to 590W by linear family extrapolation.', 320)
            out['Jinko Solar|JKM550-590-72HL4-V'] = js

    return out


def module_spec_for_row(manufacturer: str, model: str, family: str, seed_specs: Dict[str, Dict[str, Any]], source_url: str) -> Optional[Dict[str, Any]]:
    m = clean_text(model).upper()
    mf = clean_text(manufacturer)

    # Datasheet-backed seed families
    for k in (
        f'{mf}|{family}',
        'Jinko Solar|JKM550-590-72HL4-V',
        'LONGi|LR5-72HGD-560M to 590M',
        'JA Solar|JAM66D45 (DEEP BLUE 4.0 Pro)',
    ):
        if k in seed_specs:
            spec = dict(seed_specs[k])
            if not spec.get('source_url'):
                spec['source_url'] = source_url
            return spec

    # JA72 family interpolation (engineering estimate around known JA72S30 operating point)
    if m.startswith('JAM72S30'):
        return {
            'kind': 'engineering-interp',
            'pmin': 445.0,
            'pmax': 585.0,
            'voc_min': 49.2,
            'voc_max': 51.2,
            'vmp_min': 40.5,
            'vmp_max': 42.6,
            'isc_factor': 1.06,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'JA72S30 range inferred from approved family envelope + known JA72 operating point.',
        }

    if m.startswith('JAM72S40'):
        return {
            'kind': 'engineering-interp',
            'pmin': 575.0,
            'pmax': 595.0,
            'voc_min': 50.5,
            'voc_max': 52.0,
            'vmp_min': 42.5,
            'vmp_max': 44.0,
            'isc_factor': 1.055,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'JA72S40 range inferred from approved family envelope.',
        }

    # Risen family ranges from approved list envelope (engineering interpolation)
    if m.startswith('RSM156-6-'):
        return {
            'kind': 'engineering-interp',
            'pmin': 430.0,
            'pmax': 455.0,
            'voc_min': 53.5,
            'voc_max': 54.8,
            'vmp_min': 44.8,
            'vmp_max': 46.2,
            'isc_factor': 1.06,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'RSM156-6 range inferred from approved family envelope.',
        }

    if m.startswith('RSM150-8-'):
        return {
            'kind': 'engineering-interp',
            'pmin': 480.0,
            'pmax': 500.0,
            'voc_min': 50.0,
            'voc_max': 51.2,
            'vmp_min': 41.8,
            'vmp_max': 43.2,
            'isc_factor': 1.055,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'RSM150-8 range inferred from approved family envelope.',
        }

    if m.startswith('RSM110-8-'):
        return {
            'kind': 'engineering-interp',
            'pmin': 535.0,
            'pmax': 560.0,
            'voc_min': 38.8,
            'voc_max': 40.1,
            'vmp_min': 31.8,
            'vmp_max': 33.4,
            'isc_factor': 1.06,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'RSM110-8 range inferred from approved family envelope.',
        }

    if m.startswith('RSM144-9-'):
        return {
            'kind': 'engineering-interp',
            'pmin': 535.0,
            'pmax': 555.0,
            'voc_min': 49.6,
            'voc_max': 50.2,
            'vmp_min': 41.8,
            'vmp_max': 42.2,
            'isc_factor': 1.06,
            'maxSystemV': 1500,
            'source_url': source_url,
            'source_note': 'RSM144-9 range inferred from approved family envelope + known 535W operating point.',
        }

    return None


def build_pv_and_grid_catalogs(seed_rows: List[Dict[str, Any]], expanded_rows: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    seed_specs = base_seed_family_specs(seed_rows)

    modules: List[Dict[str, Any]] = []
    module_registry: List[Dict[str, Any]] = []
    module_seen = set()

    grid_inverters: List[Dict[str, Any]] = []
    grid_seen = set()

    missing_grid_kw: List[Dict[str, Any]] = []

    for row in expanded_rows:
        item_type = clean_text(row.get('item_type'))
        manufacturer = clean_text(row.get('manufacturer'), 80)
        model = clean_text(row.get('model_exact'), 140)
        family = clean_text(row.get('model_family'), 160)
        source_url = clean_text(row.get('source_url'), 300)
        source_note = clean_text(row.get('source_note'), 300)
        confidence = clean_text(row.get('confidence'), 20).lower() or 'medium'
        if not manufacturer or not model:
            continue

        if item_type == 'PV Module':
            watt = parse_module_watt(model)
            spec = module_spec_for_row(manufacturer, model, family, seed_specs, source_url)
            calc_ready = bool(spec and watt and watt > 0)
            module_registry.append({
                'manufacturer': manufacturer,
                'model': model,
                'modelFamily': family,
                'calculationReady': calc_ready,
                'sourceUrl': source_url,
                'sourceConfidence': confidence,
                'method': spec.get('kind') if spec else 'unresolved',
            })
            if not calc_ready:
                continue

            p = float(watt)
            pmin = float(spec['pmin'])
            pmax = float(spec['pmax'])
            voc = interp(float(spec['voc_min']), float(spec['voc_max']), pmin, pmax, p)
            vmp = interp(float(spec['vmp_min']), float(spec['vmp_max']), pmin, pmax, p)

            if spec.get('kind') == 'datasheet-range':
                isc = interp(float(spec['isc_min']), float(spec['isc_max']), pmin, pmax, p)
                imp = interp(float(spec['imp_min']), float(spec['imp_max']), pmin, pmax, p)
            else:
                imp = p / max(vmp, 0.01)
                isc = imp * float(spec.get('isc_factor', 1.06))

            coeff_voc, coeff_isc, coeff_pmax = coeff_defaults(manufacturer)
            sid = slug_id(manufacturer, model)
            if sid in module_seen:
                continue
            module_seen.add(sid)

            method = spec.get('kind', 'engineering-interp')
            modules.append({
                'id': sid,
                'manufacturer': manufacturer,
                'model': model,
                'Pmax': round_n(p, 3),
                'Voc': round_n(voc, 3),
                'Vmp': round_n(vmp, 3),
                'Isc': round_n(isc, 3),
                'Imp': round_n(imp, 3),
                'coeffVoc': coeff_voc,
                'coeffIsc': coeff_isc,
                'coeffPmax': coeff_pmax,
                'coeffVmp': coeff_voc,
                'coeffImp': coeff_isc,
                'NOCT': 43,
                'NMOT': 43,
                'seriesFuseA': 30 if p >= 560 else 25,
                'maxSystemV': int(spec.get('maxSystemV') or 1500),
                'datasheetUrl': clean_text(spec.get('source_url') or source_url, 300),
                'datasheetRev': 'seed-2026-04-09',
                'tolerancePlusPct': 3.0,
                'toleranceMinusPct': 0.0,
                'cells': cells_guess(model),
                'preloaded': True,
                'note': f"Catalogue-backed ({method}) from approved-list family mapping ({family}).",
                'sourceConfidence': confidence,
                'catalogueBacked': True,
            })

        elif item_type == 'PV Inverter':
            sid = slug_id(manufacturer, model)
            if sid in grid_seen:
                continue
            grid_seen.add(sid)
            ac_kw = parse_inverter_kw(model, manufacturer)
            if ac_kw is None or ac_kw <= 0:
                missing_grid_kw.append({'manufacturer': manufacturer, 'model': model})
                continue
            grid_inverters.append({
                'id': sid,
                'manufacturer': manufacturer,
                'model': model,
                'topology': 'grid-tie',
                'acRated_kW': round_n(ac_kw, 3),
                'maxDcVoc_V': None,
                'mpptMin_V': None,
                'mpptMax_V': None,
                'mpptCount': None,
                'maxCurrentPerMppt_A': None,
                'datasheetUrl': source_url,
                'datasheetRev': 'seed-2026-04-09',
                'sourceConfidence': confidence,
                'listingSource': source_note,
            })

    modules.sort(key=lambda x: (x['manufacturer'], x['model']))
    module_registry.sort(key=lambda x: (x['manufacturer'], x['model']))
    grid_inverters.sort(key=lambda x: (x['manufacturer'], x['model']))

    pv_payload = {
        'version': '2026-04-09',
        'source': 'Approved-list staging workbook + datasheet/engineering family interpolation',
        'modules': modules,
        'registry': module_registry,
    }

    grid_payload = {
        'version': '2026-04-09',
        'source': 'Approved-list staging workbook (grid inverter model registry with parsed AC ratings)',
        'inverters': grid_inverters,
    }

    return pv_payload, grid_payload, module_registry, missing_grid_kw


def hybrid_defaults(manufacturer: str, model: str, ac_kw: float) -> Dict[str, Any]:
    mf = clean_text(manufacturer).upper()
    u = clean_text(model).upper()

    bus = 48.0
    mppt_min = 150.0
    mppt_max = 550.0
    max_voc = 600.0
    mppt_count = 2
    max_cur_mppt = 16.0
    surge_factor = 2.0
    surge_s = 10.0
    max_pv = ac_kw * 1.30
    charge_a = max(60.0, ac_kw * 1000.0 / bus * 1.0)
    discharge_a = charge_a

    if 'SOFAR' in mf:
        if '3PH' in u:
            bus = 400.0
            mppt_min = 180.0
            mppt_max = 960.0
            max_voc = 1000.0
            max_cur_mppt = 26.0
            surge_factor = 1.5
            surge_s = 5.0
            max_pv = ac_kw * 1.35
            charge_a = max(20.0, ac_kw * 1000.0 / bus * 1.2)
            discharge_a = charge_a
        elif u.startswith('ESI'):
            bus = 400.0
            mppt_min = 85.0
            mppt_max = 520.0
            max_voc = 550.0
            max_cur_mppt = 16.0
            surge_factor = 1.5
            surge_s = 5.0
            max_pv = ac_kw * 1.25
            charge_a = 20.0
            discharge_a = 20.0
        elif u.startswith('HYD'):
            bus = 48.0
            mppt_min = 90.0
            mppt_max = 550.0
            max_voc = 600.0
            max_cur_mppt = 13.0
            surge_factor = 2.0
            surge_s = 10.0
            max_pv = ac_kw * 1.30
            charge_a = 100.0
            discharge_a = 100.0

    elif 'SOLIS' in mf:
        bus = 48.0
        mppt_min = 90.0
        mppt_max = 520.0
        max_voc = 600.0
        max_cur_mppt = 16.0
        surge_factor = 2.0
        surge_s = 10.0
        max_pv = ac_kw * 1.30
        charge_a = 125.0 if ac_kw >= 5 else 100.0
        discharge_a = charge_a

    elif 'DEYE' in mf:
        bus = 48.0
        surge_factor = 2.0
        surge_s = 10.0
        if 'LP3' in u:
            mppt_min = 150.0
            mppt_max = 850.0
            max_voc = 1000.0
            max_cur_mppt = 26.0
            max_pv = ac_kw * 1.50
            charge_a = max(160.0, ac_kw * 1000.0 / bus * 0.9)
            discharge_a = charge_a
        elif 'SG01' in u:
            mppt_min = 150.0
            mppt_max = 500.0
            max_voc = 550.0
            max_cur_mppt = 26.0
            max_pv = ac_kw * 1.40
            charge_a = max(190.0, ac_kw * 1000.0 / bus * 0.9)
            discharge_a = charge_a
        else:
            mppt_min = 150.0
            mppt_max = 425.0
            max_voc = 500.0
            max_cur_mppt = 13.0
            max_pv = ac_kw * 1.30
            charge_a = max(120.0, ac_kw * 1000.0 / bus * 0.95)
            discharge_a = charge_a

    elif 'KACO' in mf:
        bus = 400.0
        mppt_min = 150.0
        mppt_max = 850.0
        max_voc = 1000.0
        max_cur_mppt = 18.0
        surge_factor = 1.5
        surge_s = 5.0
        max_pv = ac_kw * 1.35
        charge_a = max(20.0, ac_kw * 1000.0 / bus * 1.2)
        discharge_a = charge_a

    return {
        'surge_kW': round_n(ac_kw * surge_factor, 3),
        'surge_s': round_n(surge_s, 3),
        'batteryBus_V': round_n(bus, 3),
        'maxCharge_A': round_n(charge_a, 3),
        'maxDischarge_A': round_n(discharge_a, 3),
        'maxPv_kW': round_n(max_pv, 3),
        'mpptMin_V': round_n(mppt_min, 3),
        'mpptMax_V': round_n(mppt_max, 3),
        'maxDcVoc_V': round_n(max_voc, 3),
        'mpptCount': int(mppt_count),
        'maxCurrentPerMppt_A': round_n(max_cur_mppt, 3),
    }


def build_hybrid_catalog(expanded_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    existing: List[Dict[str, Any]] = []
    if OUT_HYBRID.exists():
        try:
            cur = json.loads(OUT_HYBRID.read_text(encoding='utf-8'))
            existing = cur.get('inverters', []) if isinstance(cur, dict) else []
        except Exception:
            existing = []

    rows: List[Dict[str, Any]] = []
    seen = set()

    for row in expanded_rows:
        if clean_text(row.get('item_type')) != 'PV Hybrid Inverter':
            continue

        manufacturer = clean_text(row.get('manufacturer'), 80)
        model = clean_text(row.get('model_exact'), 140)
        source_url = clean_text(row.get('source_url'), 300)
        source_note = clean_text(row.get('source_note'), 300)
        confidence = clean_text(row.get('confidence'), 20).lower() or 'medium'
        if not manufacturer or not model:
            continue

        ac_kw = parse_inverter_kw(model, manufacturer)
        if ac_kw is None or ac_kw <= 0:
            continue

        item_id = slug_id(manufacturer, model)
        if item_id in seen:
            continue
        seen.add(item_id)

        defaults = hybrid_defaults(manufacturer, model, ac_kw)
        rows.append({
            'id': item_id,
            'manufacturer': manufacturer,
            'model': model,
            'acRated_kW': round_n(ac_kw, 3),
            **defaults,
            'supportedProfiles': ['offgrid', 'ceb_2025', 'leco_2025'],
            'utilityListed': {
                'ceb_2025': True,
                'leco_2025': True,
            },
            'listingSource': {
                'ceb_2025': source_note or 'Approved-list workbook 2026-04-09',
                'leco_2025': source_note or 'Approved-list workbook 2026-04-09',
            },
            'datasheetRev': 'seed-2026-04-09',
            'datasheetUrl': source_url,
            'note': f'Catalogue-backed hybrid model ({confidence} confidence).',
        })

    # Keep any existing catalog entries that are not in the approved-list build
    for old in existing:
        if not isinstance(old, dict):
            continue
        oid = clean_text(old.get('id'), 120)
        if not oid:
            oid = slug_id(clean_text(old.get('manufacturer')), clean_text(old.get('model')))
            if not oid:
                continue
            old['id'] = oid
        if oid in seen:
            continue
        seen.add(oid)
        rows.append(old)

    rows.sort(key=lambda x: (clean_text(x.get('manufacturer')), clean_text(x.get('model'))))

    return {
        'version': '2026-04-09',
        'source': 'Approved-list staging workbook for hybrid inverter coverage + preserved legacy entries',
        'inverters': rows,
    }


def build_catalogs() -> None:
    seed_rows = read_sheet_rows(SEED_XLSX, 'PV_Modules')
    expanded_rows = read_sheet_rows(EXPANDED_XLSX, 'Models_Staging_Exact')

    pv_payload, grid_payload, module_registry, missing_grid_kw = build_pv_and_grid_catalogs(seed_rows, expanded_rows)
    hybrid_payload = build_hybrid_catalog(expanded_rows)

    OUT_PV.write_text(json.dumps(pv_payload, indent=2, ensure_ascii=True) + '\n', encoding='utf-8')
    OUT_GRID.write_text(json.dumps(grid_payload, indent=2, ensure_ascii=True) + '\n', encoding='utf-8')
    OUT_HYBRID.write_text(json.dumps(hybrid_payload, indent=2, ensure_ascii=True) + '\n', encoding='utf-8')

    total_calc_ready = sum(1 for x in module_registry if x.get('calculationReady'))
    print(f'Wrote {OUT_PV} ({total_calc_ready} calc-ready modules, {len(module_registry)} registry rows)')
    print(f'Wrote {OUT_GRID} ({len(grid_payload.get("inverters", []))} grid inverters)')
    print(f'Wrote {OUT_HYBRID} ({len(hybrid_payload.get("inverters", []))} hybrid inverters)')
    if missing_grid_kw:
        print(f'Grid models without parsed kW: {len(missing_grid_kw)}')
        for miss in missing_grid_kw[:30]:
            print('  -', miss['manufacturer'], miss['model'])


if __name__ == '__main__':
    build_catalogs()




